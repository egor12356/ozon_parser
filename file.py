import openpyxl
from openpyxl.styles import PatternFill
from typing import List

from typing_class import ItemInput


def save_one_item(dict_input, number_row, file_name):
    """Добавляем все товары в файл"""
    wb = openpyxl.load_workbook(file_name)
    my_sheet = wb.active

    art = dict_input.get('art')
    ean = dict_input.get('ean')
    name_seller = dict_input.get('name_seller')
    brand = dict_input.get('brand')
    name = dict_input.get('name')
    price = dict_input.get('price')
    sale = dict_input.get('sale')

    # Это товар поставщика, красим его в цвет
    if dict_input.get('not_save'):
        color = 'ffff00'

        # Записываем баркод
        my_sheet[f'A{number_row}'] = ean
        x3 = my_sheet[f'A{number_row}']
        x3.fill = PatternFill(fill_type='solid', start_color=color)

        # Записываем бренд
        my_sheet[f'B{number_row}'] = brand
        x2 = my_sheet[f'B{number_row}']
        x2.fill = PatternFill(fill_type='solid', start_color=color)

        # Записываем название книги
        my_sheet[f'C{number_row}'] = name
        x5 = my_sheet[f'C{number_row}']
        x5.fill = PatternFill(fill_type='solid', start_color=color)

        # Записываем артикул
        my_sheet[f'D{number_row}'] = art
        x1 = my_sheet[f'D{number_row}']
        x1.fill = PatternFill(fill_type='solid', start_color=color)

        # Записываем цену
        my_sheet[f'E{number_row}'] = price
        x6 = my_sheet[f'E{number_row}']
        x6.fill = PatternFill(fill_type='solid', start_color=color)

        # Записываем продавец
        my_sheet[f'F{number_row}'] = name_seller
        x4 = my_sheet[f'F{number_row}']
        x4.fill = PatternFill(fill_type='solid', start_color=color)
        # print(1)


        # Записываем sale
        if sale:
            my_sheet[f'G{number_row}'] = sale
            x5 = my_sheet[f'F{number_row}']
            x5.fill = PatternFill(fill_type='solid', start_color=color)
            # print(1)

    else:
        # Записываем баркод
        my_sheet[f'A{number_row}'] = ean

        # Записываем бренд (издательство)
        my_sheet[f'B{number_row}'] = brand

        # Записываем название книги
        my_sheet[f'C{number_row}'] = name

        # Записываем артикул
        my_sheet[f'D{number_row}'] = art

        # Записываем цену
        my_sheet[f'E{number_row}'] = price

        # Записываем название продавца
        my_sheet[f'F{number_row}'] = name_seller


    number_row += 1


    wb.save(file_name)
    wb.close()

    return number_row


def read_file() -> List[ItemInput]:
    my_path = "template.xlsx"
    wb = openpyxl.load_workbook(my_path)
    my_sheet = wb.active

    # Число строчек на странице
    max_row = my_sheet.max_row

    arr_input = []

    for number_row in range(2, max_row + 1):

        code = my_sheet[f'A{number_row}'].value
        name = my_sheet[f'B{number_row}'].value.replace(' ', '')
        isbn = my_sheet[f'C{number_row}'].value.replace(' ', '')
        ean = my_sheet[f'D{number_row}'].value
        year = my_sheet[f'E{number_row}'].value
        publishing_house = my_sheet[f'D{number_row}'].value
        item_input = ItemInput(code=code, isbn=isbn, ean=ean, publishing_house=publishing_house, year=year)

        # Остановка, если нет кода
        if not code:
            break
        arr_input.append(item_input)
    # print(arr_input)
    wb.close()
    return arr_input

if __name__ == '__main__':
    print(read_file())